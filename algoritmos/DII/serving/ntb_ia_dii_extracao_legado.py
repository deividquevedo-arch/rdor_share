# Databricks notebook source
# MAGIC %md
# MAGIC # Instalação de libs

# COMMAND ----------

# DBTITLE 1,pip install
!pip install -q openpyxl xlsxwriter

# COMMAND ----------

# DBTITLE 1,restartPython
dbutils.library.restartPython()

# COMMAND ----------

# MAGIC %md
# MAGIC

# COMMAND ----------

# MAGIC %md
# MAGIC #Imports

# COMMAND ----------

# DBTITLE 1,Import
import os
import pandas as pd
import openpyxl
import copy
import json
import requests
import subprocess

from dataclasses import dataclass, asdict
from datetime import date
from pathlib import Path

# COMMAND ----------

# DBTITLE 1,pd.set_option
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.max_colwidth', 230)

# COMMAND ----------

# MAGIC %md
# MAGIC # Parâmetros do notebook

# COMMAND ----------

# dbutils.widgets.removeAll()

# COMMAND ----------

# DBTITLE 1,Ambiente
dbutils.widgets.text("environment", "dev", "Environment")
environment = dbutils.widgets.get("environment")
print("Environment:", environment)

# COMMAND ----------

# DBTITLE 1,ID Projeto
dbutils.widgets.text("id_projeto", "dii", "ID Projeto")
id_projeto = dbutils.widgets.get("id_projeto")
print("id_projeto:", id_projeto)

# COMMAND ----------

# DBTITLE 1,Data Execução Modelo
from datetime import datetime

dbutils.widgets.text("data_execucao_modelo", "", "Data Execução Modelo")
data_execucao_modelo = dbutils.widgets.get("data_execucao_modelo")
if data_execucao_modelo == "":
    data_execucao_modelo = datetime.now().strftime("%Y-%m-%d")

print(f"Data Execução Modelo: {data_execucao_modelo}")

# COMMAND ----------

# DBTITLE 1,Catalog
dbutils.widgets.text("catalog", "ia", "Catalog")
catalog_name = dbutils.widgets.get("catalog")
print(f"catalog_name: {catalog_name}")

# COMMAND ----------

# DBTITLE 1,Prefixo Tabelas
environment_tbl = "" if environment in ["hml", "prd"] else f"{environment}_"
print("environment_tbl:", environment_tbl)

# COMMAND ----------

# DBTITLE 1,Define variável que será injetada no nome das tabelas
if environment in ["hml", "prd", "dev"]:
    main_catalog = catalog_name + "."
else:
    main_catalog = catalog_name + "." + f"{environment_tbl}"

print("main_catalog:", main_catalog)

# COMMAND ----------

# DBTITLE 1,Define variáveis com nomes das tabelas
tbl_saida = f"{main_catalog}tb_diamond_mod_{id_projeto}_saida"
print(tbl_saida)

# COMMAND ----------

# DBTITLE 1,Define o caminho principal da pasta remota
root_remote_path = f"/mnt/trusted/datalake/ia/projetos/{id_projeto}/"
print(root_remote_path)

# COMMAND ----------

# DBTITLE 1,Define o caminho principal da pasta remota de configs
root_remote_config_path = f"{root_remote_path}config/{environment}/"
print(root_remote_config_path)

# COMMAND ----------

# DBTITLE 1,Define o caminho principal da pasta remota dos dados
root_remote_data_path = f"{root_remote_path}data/{environment}/envio/"
print(root_remote_data_path)

# COMMAND ----------

# DBTITLE 1,Define o caminho mensal da pasta remota
year, month, day = data_execucao_modelo.split("-")
remote_path = f"{root_remote_data_path}{year}/{month}/"
print(remote_path)

# COMMAND ----------

# DBTITLE 1,Constantes
# IDADE_LIMITE = 75

# COMMAND ----------

# DBTITLE 1,Obtém o caminho da pasta atual
# shell_command = "pwd"
# output = subprocess.check_output(shell_command, shell=True)
# current_folder = os.path.join(output.decode().strip(), id_projeto) + "/"

# current_folder = os.path.join(os.getcwd(), id_projeto) + "/"

current_folder = os.path.join("/tmp", id_projeto) + "/"

Path(current_folder).mkdir(parents=True, exist_ok=True)

print(current_folder)

# COMMAND ----------

# DBTITLE 1,Exebe o conteúdo da pasta atual
ls -lha {current_folder}

# COMMAND ----------

# MAGIC %md
# MAGIC #Funções auxiliares

# COMMAND ----------

# DBTITLE 1,get_data
def get_data(unidades):
  df = spark.sql(f"""
      with cte_input as (
            select
                  dataExecucaoModelo,
                  fl_relevante as flgRelevante,
                  exm_laudo_achados as laudoExameAchado,
                  exm_laudo_resultado as laudoExameResultado,
                  exm_laudo_texto_tratado as laudoExameTratado,
                  exm_laudo_texto as laudoExameOriginal,
                  exm_laudo_dataliber as laudoExameDataLiberacao,
                  exm_status as exameStatus,
                  exm_an as exameAn,
                  exm_tipo as tipoExame,
                  exm_titulo as exameTitulo,
                  exm_mod as exameMod,
                  exm_data as exameData,
                  gold_corporativo.default.rdsl_decrypt(telefonePaciente,0) as telefonePaciente,
                  telefonePacienteDDD as telefonePacienteDDD,
                  idade_paciente as idadePaciente,
                  pct_datanasc as dataNascimentoPaciente,
                  pct_sexo as sexoPaciente,
                  gold_corporativo.default.rdsl_decrypt(pct_cpf,1) as cpfPaciente,
                  gold_corporativo.default.rdsl_decrypt(pct_nome,1) as nomePaciente,
                  unidade as nomeHospital,
                  idunidade as idHospital,
                  id_pct as idPaciente,
                  record_id as recordId,
                  nme_regional_hospital as ufRegional
            from {tbl_saida}
            where dataExecucaoModelo = date('{data_execucao_modelo}')
             and fl_relevante = 1
             and idunidade in ({unidades})
      )
      ,cte as (
            select
                dataExecucaoModelo,
                ufRegional,
                nomeHospital,
                nomePaciente,
                cpfPaciente, 
                sexoPaciente,
                idadePaciente,
                telefonePaciente, 
                exameAn,
                laudoExameDataLiberacao,
                tipoExame,
                laudoExameOriginal,
                laudoExameResultado,
                laudoExameAchado,
                exameMod,
                exameData,
                exameTitulo,

                  '' as achado_relevante,
                  '' as oncologia, 
                  '' as comentarios,
                  '' as cadatro_convenio,
                  '' as cadastro_plano,
                  '' as data_liberacao_exame,
                  '' as material,
                  '' as data_1_contato_MSG,
                  '' as data_2_contato_tel,
                  '' as data_3_contato_MSG,
                  '' as paciente_navegado,
                  '' as data_admissao_navegacao,
                  '' as mes,
                  '' as ano,
                  '' as data_alta_navegacao,
                  '' as motivo_alta_navegacao,
                  '' as diagnostico,
                  '' as status_1_consulta_especialista,
                  '' as data_1_consulta_especialista,
                  '' as nome_medico,
                  '' as unidade_gendamento,
                  '' as tipo_tratamento,
                  '' as cirugia_solicitada,
                  '' as data_pedido,
                  '' as data_cirugia,
                  '' as local_cirurgia,
                  '' as observacoes,

                  "" as achadoRelevante,
                  "" as linhaCuidado,
                  "" as destino,
                  "" as prioridade,

                  case 
                        when trim(upper(ufRegional)) = 'RJ' then 1
                        when trim(upper(ufRegional)) = 'SP' then 2
                        when trim(upper(ufRegional)) = 'PB' then 3
                        else 4
                  end as regionalHospitalOrder
            from cte_input
      )
      select * except(regionalHospitalOrder)
      from cte
      order by
        regionalHospitalOrder,
        nomePaciente,
        exameData

  """).toPandas()

  return df

# COMMAND ----------

# DBTITLE 1,Função get_cols
def get_cols():
    """Dicionário de/para com os nomes das colunas de saída"""
    dic_col_names = {
        "ufRegional" : "Regional",
        "nomeHospital": "Unidade", 
        "nomePaciente": "Nome Paciente",
        "cpfPaciente" : "CPF",
        "sexoPaciente" : "Sexo",
        "idadePaciente": "Idade",
        "telefonePaciente": "Telefone",
        "exameAn": "Codigo AN",
        "laudoExameDataLiberacao": "Data de Liberacao do Laudo",
        "tipoExame": "Origem Exame",
        "exameMod" : "Modalidade",
        "exameTitulo": "Tipo Exame",
        "laudoExameOriginal": "Laudo", 
        "laudoExameResultado": "Resultado", 
        "laudoExameAchado" : "Achados",
        "dataExecucaoModelo": "Data de Execucao do Modelo",

        "achado_relevante" : "Achado Relevante",
        "oncologia" : "Oncologia", 
        "comentarios" : "Comentarios",
        "cadatro_convenio" : "Cadastro Convenio",
        "cadastro_plano" : "Cadastro Plano",
        "data_liberacao_exame" : "Data de Liberacao de Exame",
        "material" : "Material",
        "data_1_contato_MSG" : "Data 1º Contato | MSG",
        "data_2_contato_tel" : "Data 2º Contato | Tel",
        "data_3_contato_MSG" : "Data 3º Contato | MSG",
        "paciente_navegado" : "Paciente Navegado",
        "data_admissao_navegacao" : "Data de Admissao na Navegacao",
        "mes" : "Mes",
        "ano" : "Ano",
        "data_alta_navegacao" : "Data da Alta da Navegacao",
        "motivo_alta_navegacao" : "Motivo da Alta da Navegacao",
        "diagnostico" : "Diagnostico",
        "status_1_consulta_especialista" : "Status da 1º Consulta com o Especialista",
        "data_1_consulta_especialista" : "Data 1º Consulta com o Especialista",
        "nome_medico" : "Nome do Medico",
        "unidade_gendamento" : "Unidade de Agendamento",
        "tipo_tratamento" : "Tipo de Tratamento",
        "cirugia_solicitada" : "Cirugia Solicitada",
        "data_pedido" : "Data do Pedido",
        "data_cirugia" : "Data da Cirugia",
        "local_cirurgia" : "Local da Cirurgia",
        "observacoes" : "Observacoes",

        # "linhaCuidado": "Linha de Cuidado", removido apos conversa com Natan em 2026-01-23
    }

    return dic_col_names

# COMMAND ----------

# DBTITLE 1,Função export_to_excel
def export_to_excel(df, dic_col_names, unidade):
    file_name = f"{id_projeto}_{unidade}_{data_execucao_modelo}.xlsx"
    local_file = f"{current_folder}{file_name}"
    remote_file = f"{remote_path}{file_name}"

    cols = [col for col in dic_col_names.keys()]
    df[cols].to_excel(local_file, index=False, engine='xlsxwriter')

    return {
        "unidade": unidade,
        "nomeArquivo": file_name,
        "arquivoLocal": local_file,
        "arquivoRemoto": remote_file,
        "caminhoRemoto": remote_path.replace("/mnt", ""),
        "registros": len(df.index),
        "dataProcessamento": data_execucao_modelo,
        "dataProcessamentoFormatada": datetime.strptime(data_execucao_modelo, "%Y-%m-%d").strftime("%d/%m/%Y"),
    }

# COMMAND ----------

# DBTITLE 1,Função get_cols_config
def get_cols_config():
    """Define tamanho fixo para algumas colunas. ( -1 = Coluna oculta )"""
    dic_col_fixed_width = {
        "Data de Execucao do Modelo": -1,
        "Laudo": 150,
        "Achados": 80,
        "Resultado": 130,
        "Observações": 30,
        "Achado Relevante": 37,
        "Linha de Cuidado": 27,
        "Destino": 80,
        
    }
    return dic_col_fixed_width

# COMMAND ----------

from openpyxl.worksheet.datavalidation import DataValidation

def create_list_validation(wb, sheet_name, list_title, values, col_letter, start_row=2, hide_sheet=True):
    """
    Cria/atualiza uma lista em uma aba e devolve um DataValidation que aponta para o range dessa lista.
    - values: lista de strings (1 por linha)
    - col_letter: letra da coluna onde a lista vai ficar (ex: "A", "B")
    """
    if sheet_name in wb.sheetnames:
        ws_list = wb[sheet_name]
    else:
        ws_list = wb.create_sheet(sheet_name)

    # cabeçalho
    ws_list[f"{col_letter}1"].value = list_title

    # limpa valores antigos (opcional simples: sobrescreve)
    for i, v in enumerate(values, start=start_row):
        ws_list[f"{col_letter}{i}"].value = v

    last_row = start_row + len(values) - 1

    if hide_sheet:
        ws_list.sheet_state = "hidden"

    dv = DataValidation(
        type="list",
        formula1=f"={sheet_name}!${col_letter}${start_row}:${col_letter}${last_row}",
        allow_blank=True
    )
    return dv

# COMMAND ----------

# DBTITLE 1,Função format_excel
def format_excel(full_file_name, dic_col_names, dic_col_fixed_width):
    import math
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.load_workbook(full_file_name)
    sheet = wb.active

    dic_col_size = {}

    data_validation_error_title = "Valor Inválido"
    data_validation_error_message = "Este valor não corresponde às restrições de valições de dados definidas para esta célula."

    ## config select box ↓↓↓↓↓↓⚠️⚠️⚠️⚠️⚠️

    dv_achado = DataValidation(type="list", formula1='"1 - Sim (Tem DII),2 - Não"', allow_blank=True)
    dv_achado.showInputMessage = True
    dv_achado.showErrorMessage = True
    dv_achado.error = data_validation_error_message
    dv_achado.errorTitle = data_validation_error_title

    dv_oncologia = DataValidation(type="list", formula1='"1 - Oncologico,2 - Não Oncologico"', allow_blank=True)
    dv_oncologia.showInputMessage = True
    dv_oncologia.showErrorMessage = True
    dv_oncologia.error = data_validation_error_message
    dv_oncologia.errorTitle = data_validation_error_title

    convenios = ["1 - A + SAUDE","2 - AASMED","3 - ABAS","4 - ABERTTA SAUDE","5 - ABMED","6 - ADVANCE","7 - AERONAUTICA","8 - AETNA INTERNACIONAL (EURO CENTER)","9 - AFABB","10 - AFFEGO","11 - AFRAFEP","12 - AFRESP","13 - AFUSE","14 - AGROS","15 - ALEXION","16 - ALLIANZ","17 - AMAFRE RJ","18 - AMAFRESP","19 - AMAGIS","20 - AMAI","21 - AMBEP","22 - AMEPE CAMPE","23 - AMEX SAUDE","24 - AMHA SAUDE","25 - AMMP","26 - ANADEM","27 - ANAFE SAUDE","28 - APCEF","29 - APOSTOLAS DO SAGRADO CORACAO DE JESUS","30 - APUB","31 - ARCELOR MITTAL","32 - ARTEB","33 - ASFAL","34 - ASFEB","35 - ASPFEM","36 - ASSEC SAUDE","37 - ASSEFAZ","38 - ASSEM","39 - ASSIM SAUDE","40 - ATIVIA","41 - AVAMP","42 - AWP SERVICE BRASIL","43 - BANCO CENTRAL","44 - BB SAUDE","45 - BEACH PARK","46 - BEN+","47 - BEST DOCTOR","48 - BIO SAUDE","49 - BRADESCO","50 - BRADESCO FUNCIONARIOS","51 - BRASKEM","52 - BUPA","53 - CAASP","54 - CABERJ","55 - CABESP","56 - CAFAZ","57 - CAMARA DOS DEPUTADOS","58 - CAMARJ","59 - CAMED SAUDE","60 - CAMPE","61 - CAMPERJ","62 - CAPESESP","63 - CARE PLUS","64 - CARTAO EXECUTIVO","65 - CASEMBRAPA","66 - CASF","67 - CASHSERVICE","68 - CASSE","69 - CASSEB","70 - CASSEMS","71 - CASSI","72 - CASSIND","73 - CASU","74 - CAURJ","75 - CBC","76 - CBMERJ","77 - CELPE","78 - CEMEC","79 - CEMERU - OPLAN","80 - CEMIG SAUDE","81 - CENIBRA","82 - CENTRAL NACIONAL UNIMED","83 - CET","84 - CETESB","85 - CFMG","86 - CINDACTA (FORCA AEREA)","87 - CNEN/CDTN","88 - CNEN/IPEN","89 - CODEVASF","90 - COMPESA SAUDE","91 - COMSEDER","92 - CONAB","93 - CONEGAS SANTO AGOSTINHO","94 - CONGREGACAO CATEQUISTA DO SAGRADO CORACAO","95 - CONGREGACAO IMACULADA CONCEICAO","96 - CONGREGACAO NOSSA SENHORA DE SION","97 - COPASS","98 - COPEL","99 - CRUZ AZUL","100 - DESBAN","101 - ECONOMUS","102 - EFHARMA","103 - ELETROS SAUDE","104 - EMBAIXADA AMERICANA","105 - EMBRATEL (GRUPO EMBRATEL)","106 - EMBRATEL (PAME)","107 - ENZIMAS","108 - EPHARMA","109 - ESPORTE CLUBE PINHEIROS","110 - ESSENTIAL STELARA","111 - E-VIDA","112 - EXMED","113 - FACEB","114 - FACHESF","115 - FAMED SAUDE","116 - FAMEH","117 - FAPES - BNDES","118 - FASCAL","119 - FAVC","120 - FCA SAUDE (FIAT)","121 - FERINJECT TAKEDA","122 - FIOSAUDE","123 - FISCO SAUDE","124 - FUNASA","125 - FUNDACAO FIAT","126 - FUNDACAO LIBERTAS","127 - FUNDACAO SAUDE ITAU","128 - FUNDAFFEMG SAUDE","129 - FUSEX","130 - FUSMA","131 - GACC SUS","132 - GAMA SAUDE","133 - GEAP","134 - GENTE SEGURADORA","135 - GH SAUDE","136 - GOLDEN CROSS","137 - GRUPO AMIL","138 - GRUPO SAO JOSE SAUDE","139 - HAPVIDA","140 - HFA","141 - HFAB","142 - HGU","143 - HOSPITAL NAVAL","144 - HOSPITAU","145 - HUMANIZAR","146 - IBPCLIN","147 - IDOR","148 - IMCAS","149 - IMPCG","150 - IMTEP SAUDE EMPRESARIAL","151 - IN HEALTH","152 - INAS","153 - INB","154 - INSPETORIA NOSSA SENHORA APARECIDA","155 - INSTITUTO DAS FILHAS DE NOSSA SENHORA DA MISERICODIA","156 - INSTITUTO DE MISERICORDIA","157 - INSTITUTO RELIGIOSO BARBARA MAIX - IRBM","158 - INSTITUTO SOLIDARE II","159 - INTEGRA MEDICAL","160 - INTEGRAL SAUDE","161 - INTERASSIST","162 - INTERCLINICAS","163 - IPADE","164 - IPALERJ","165 - IPES","166 - IPM","167 - IPSEMG","168 - IPSM","169 - IRH - SASSEPE","170 - IRMAS APOSTOLINAS","171 - IRMAS SAO JOSE","172 - ITACA","173 - JUDICIMED","174 - KIPP","175 - LIFE EMPRESARIAL","176 - MAHLE METAL LEVE","177 - MAPFRE","178 - MARINHA","179 - MAX SAUDE","180 - MEDIPLAN","181 - MEDISERVICE","182 - MEDSENIOR","183 - MERCEDES BENZ","184 - METRUS","185 - MINERACAO CARAIBA","186 - MINISTERIO PUBLICO DE SP","187 - MIRABILANDIA","188 - MUTUA","189 - NIPOMED","190 - NORDESTE SAUDE","191 - NOTRE DAME INTERMEDICA","192 - NOVARTIS","193 - NUCLEP","194 - OMINT","195 - OPEN LINE (CARTAO DE DESCONTO)","196 - OPV","197 - OSSEL","198 - PARANA CLINICAS","199 - PARK SHOPPING","200 - PARQUE DA MONICA","201 - PARTICULAR","202 - PASA","203 - PIA SOCIEDADE FILHAS DE SAO PAULO (PAULINAS)","204 - PLAMED","205 - PLAMER","206 - PLAN ASSISTE","207 - PLANSERV","208 - PLENA SAUDE","209 - POLICIA MILITAR","210 - PORTO SEGURO","211 - POSTAL SAUDE","212 - PREMIUM SAUDE","213 - PREVENT SENIOR","214 - PROASA","215 - PROMED","216 - PROMEDICA","217 - PRO-SOCIAL","218 - REAL GRANDEZA","219 - RENASCER-AMAVIDA","220 - S.C. CORINTHIANS","221 - SAAE","222 - SAMP/AGMP","223 - SANEPAR","224 - SANTA CASA DE MAUA","225 - SANTA CASA SAO JOSE DOS CAMPOS","226 - SANTAMALIA","227 - SAO FRANCISCO SAUDE","228 - SAO FRANSCISCO VIDA","229 - SAO PAULO FUTEBOL","230 - SAUDE BRB","231 - SAUDE CAIXA","232 - SAUDE PETROBRAS","233 - SAUDE RECIFE","234 - SAUDE SIM","235 - SAUDE SISTEMA","236 - SBC SAUDE","237 - SEGURADORAS INTERNACIONAIS","238 - SEGUROS UNIMED","239 - SENADO FEDERAL","240 - SENDAS","241 - SEPACO","242 - SERPRAM","243 - SERPRO","244 - SHOPPING GUARARAPES","245 - SHOPPING PLAZA","246 - SHOPPING RIO MAR","247 - SHOPPING TACARUNA","248 - SMILE","249 - SPA SAUDE","250 - STELLANTIS","251 - STF","252 - STF MED","253 - STJ","254 - STM","255 - SULAMERICA","256 - SULAMERICA FUNCIONARIOS","257 - TELOS","258 - TJDFT","259 - TOTAL MEDCARE","260 - TOYOTA","261 - TRANSMONTANO","262 - TRANSPORTES AEREOS PORTUGUESES (TAP)","263 - TRE","264 - TRF","265 - TRIBUNAL DE JUSTICA ESTADO DE SP","266 - TRT","267 - TST","268 - UFMS","269 - UHS BRAZIL GESTAO DE SAUDE","270 - UNAFISCO SAUDE","271 - UNIAO MEDICA","272 - UNIMED BH","273 - UNIMED CACAPAVA","274 - UNIMED CAMPO GRANDE","275 - UNIMED EXTREMO SUL","276 - UNIMED FAMA","277 - UNIMED FEDERACAO","278 - UNIMED FORTALEZA","279 - UNIMED MACEIO","280 - UNIMED NORTE NORDESTE","281 - UNIMED NOVA IGUACU","282 - UNIMED OS BANDEIRANTES","283 - UNIMED RECIFE","284 - UNIMED RESENDE","285 - UNIMED RIO","286 - UNIMED SAO JOSE DOS CAMPOS","287 - UNIMED TAQUARI","288 - UNIMED VALE SAO FRANCISCO","289 - UNISAUDE","290 - UPR","291 - URMES","292 - USISAUDE","293 - VALE","294 - VALE SAUDE","295 - VALE SAUDE SEMPRE","296 - VERITAS","297 - VITALLIS SAUDE","298 - VIVA MAIS","299 - VIVEST","300 - WORLD ASSIST","301 - PASA GTS","302 - FUNDACAO CESP"
    ]
    dv_cadastro_convenio = create_list_validation(
    wb=wb,
    sheet_name="Listas",
    list_title="Cadastro Convenio",
    values=convenios,
    col_letter="A",   # coluna A para convênios
    )

    dv_paciente_navegado = DataValidation(type="list", formula1='"1 - Sim,2 - Não"', allow_blank=True)
    dv_paciente_navegado.showInputMessage = True
    dv_paciente_navegado.showErrorMessage = True
    dv_paciente_navegado.error = data_validation_error_message
    dv_paciente_navegado.errorTitle = data_validation_error_title

    motivos = ["1 - Alta Clínica","2 - Ausência de telefone","3 - Captado","4 - Convênio não elegível","5 - Descredenciamento de convênio","6 - Desistência","7 - Dificuldade de contato","8 - Em abordagem","9 - Em tratamento com QT | Rádio","10 - Já em acompanhamento no CEMED","11 - Já está sendo navegado","12 - Já operado ","13 - Material Peça","14 - Óbito","15 - Possui Médico Externo","16 - Recusa","17 - Transferência para outro serviço","18 - Tratamento no SUS"
    ]
    dv_motivo_alta_navegacao = create_list_validation(
    wb=wb,
    sheet_name="Listas",
    list_title="Motivo da Alta da Navegacao",
    values=motivos,
    col_letter="B",   # coluna A para convênios
    )

    dv_status_primeira_consula_especialista = DataValidation(type="list", formula1='"1 - Agendado,2 - Aguardando Agendamento,3 - Realizado,4 - Reagendado"', allow_blank=True)
    dv_status_primeira_consula_especialista.showInputMessage = True
    dv_status_primeira_consula_especialista.showErrorMessage = True
    dv_status_primeira_consula_especialista.error = data_validation_error_message
    dv_status_primeira_consula_especialista.errorTitle = data_validation_error_title

    unidades_agendamento = ["1 - Alphaville","2 - Anália Franco","3 - Antonio Afonso (Jacareí)","4 - Assunção","5 - Bartira","6 - Brasil","7 - Brasil Mauá (América Mauá)","8 - Campinas","9 - Central do Tatuapé (Aviccena)","10 - Central Leste (Guaianases)","11 - Central Oeste (Alphamed)","12 - Central Sul (Serra Mayor)","13 - Guarulhos","14 - Ifor","15 - Itaim","16 - Jabaquara","17 - Jabaquara - São Remo ","18 - Morumbi","19 - Novo Atibaia","20 - Orthoservice","21 - Osasco (Sino Brasileiro)","22 - Policlínica Taboão","23 - Ribeirão Pires","24 - Santa Isabel","25 - São Caetano","26 - Vila Nova Star","27 - Villa Lobos","28 - Vivalle","29 - Badim","30 - Balbino","31 - Bangu","32 - Barra D'Or","33 - Caxias D'Or","34 - Clínica São Vicente","35 - Copa D'Or","36 - Copa Star","37 - Gloria D'Or ","38 - Niterói D'Or","39 - Norte D'Or","40 - Oeste D'Or","41 - Quinta D'Or","42 - Rio Barra (Rio Mar)","43 - Rios D'Or","44 - Samer","45 - Macaé D'Or","46 - São Rafael","47 - Aliança","48 - Cardiopulmonar","49 - Memorial São José","50 - Esperança Olinda","51 - Esperança Recife","52 - São Marcos","53 - Santa Luzia","54 - Santa Helena","55 - DF Star ","56 - Clinica Sorocaba"
    ]
    dv_unidade_agendamento = create_list_validation(
    wb=wb,
    sheet_name="Listas",
    list_title="Unidade de Agendamento",
    values=unidades_agendamento,
    col_letter="C",   # coluna A para convênios
    )

    dv_tipo_tratamento = DataValidation(type="list", formula1='"1 - Clínico,2 - Cirúrgico"', allow_blank=True)
    dv_tipo_tratamento.showInputMessage = True
    dv_tipo_tratamento.showErrorMessage = True
    dv_tipo_tratamento.error = data_validation_error_message
    dv_tipo_tratamento.errorTitle = data_validation_error_title

    dv_cirurgia_solicitada = DataValidation(type="list", formula1='"1 - Sim,2 - Não"', allow_blank=True)
    dv_cirurgia_solicitada.showInputMessage = True
    dv_cirurgia_solicitada.showErrorMessage = True
    dv_cirurgia_solicitada.error = data_validation_error_message
    dv_cirurgia_solicitada.errorTitle = data_validation_error_title

    unidade_cirurgia = ["1 - Alphaville","2 - Anália Franco","3 - Antonio Afonso (Jacareí)","4 - Assunção","5 - Bartira","6 - Brasil","7 - Brasil Mauá (América Mauá)","8 - Campinas","9 - Central do Tatuapé (Aviccena)","10 - Central Leste (Guaianases)","11 - Central Oeste (Alphamed)","12 - Central Sul (Serra Mayor)","13 - Guarulhos","14 - Ifor","15 - Itaim","16 - Jabaquara","17 - Jabaquara - São Remo ","18 - Morumbi","19 - Novo Atibaia","20 - Orthoservice","21 - Osasco (Sino Brasileiro)","22 - Policlínica Taboão","23 - Ribeirão Pires","24 - Santa Isabel","25 - São Caetano","26 - Vila Nova Star","27 - Villa Lobos","28 - Vivalle","29 - Badim","30 - Balbino","31 - Bangu","32 - Barra D'Or","33 - Caxias D'Or","34 - Clínica São Vicente","35 - Copa D'Or","36 - Copa Star","37 - Gloria D'Or ","38 - Niterói D'Or","39 - Norte D'Or","40 - Oeste D'Or","41 - Quinta D'Or","42 - Rio Barra (Rio Mar)","43 - Rios D'Or","44 - Samer","45 - Macaé D'Or","46 - São Rafael","47 - Aliança","48 - Cardiopulmonar","49 - Memorial São José","50 - Esperança Olinda","51 - Esperança Recife","52 - São Marcos","53 - Santa Luzia","54 - Santa Helena","55 - DF Star","56 - Clinica Sorocaba"
    ]
    dv_unidade_cirurgia = create_list_validation(
    wb=wb,
    sheet_name="Listas",
    list_title="Local da Cirurgia",
    values=unidade_cirurgia,
    col_letter="D",   # coluna A para convênios
    )

    ## config select box ↑↑↑↑↑↑⚠️⚠️⚠️⚠️⚠️

    # col_idade = ""

    wrap_text = [
        "Laudo",
        "Tipo Exame",
        "Nome Hospital",
    ]

    h_align = [
        "Regional",
        "Idade",
        "Telefone",
        "Data de Liberacao do Laudo",
        "Origem Exame",
        "Modalidade",
        "Achado Relevante",
    ]

    for cols in sheet.iter_cols():
        for cell in cols:
            # Renomeia as colunas
            if cell.row == 1:
                cell.value = dic_col_names.get(cell.value, cell.value)

            column_letter = cell.column_letter
            len_value = len(str(cell.value))
            
            if len_value > dic_col_size.get(column_letter, 0):
                dic_col_size[column_letter] = len_value
            
            ## config select box ↓↓↓↓↓↓⚠️⚠️⚠️⚠️⚠️

            if cols[0].value == "Achado Relevante" and cell.row > 1:
                dv_achado.add(cell)

            if cols[0].value == "Oncologia" and cell.row > 1:
                dv_oncologia.add(cell)

            if cols[0].value == "Cadastro Convenio" and cell.row > 1:
                dv_cadastro_convenio.add(cell)

            if cols[0].value == "Paciente Navegado" and cell.row > 1:
                dv_paciente_navegado.add(cell)

            if cols[0].value == "Motivo da Alta da Navegacao" and cell.row > 1:
                dv_motivo_alta_navegacao.add(cell)

            if cols[0].value == "Status da 1º Consulta com o Especialista" and cell.row > 1:
                dv_status_primeira_consula_especialista.add(cell)

            if cols[0].value == "Unidade de Agendamento" and cell.row > 1:
                dv_unidade_agendamento.add(cell)

            if cols[0].value == "Tipo de Tratamento" and cell.row > 1:
                dv_tipo_tratamento.add(cell)

            if cols[0].value == "Cirurgia Solicitada" and cell.row > 1:
                dv_cirurgia_solicitada.add(cell)

            if cols[0].value == "Local da Cirurgia" and cell.row > 1:
                dv_unidade_cirurgia.add(cell)

            ## config select box ↑↑↑↑↑↑⚠️⚠️⚠️⚠️⚠️

            alignment = copy.copy(cell.alignment)
            alignment.vertical = "center"

            if cell.row > 1 and cols[0].value in wrap_text:
                alignment.wrapText = True

            if cell.row > 1 and cols[0].value in h_align:
                alignment.horizontal = "center"
        
            cell.alignment = alignment

    # Ajusta a largura das colunas
    for k, v in dic_col_size.items():
        adjusted_width = dic_col_fixed_width.get(sheet[f"{k}1"].value)

        if adjusted_width is None:
            adjusted_width = int((v + 2))

        if adjusted_width == -1:
            sheet.column_dimensions[k].hidden = True
        else:
            sheet.column_dimensions[k].width = adjusted_width

    sheet.add_data_validation(dv_achado)
    sheet.add_data_validation(dv_oncologia)
    sheet.add_data_validation(dv_cadastro_convenio)
    sheet.add_data_validation(dv_paciente_navegado)
    sheet.add_data_validation(dv_motivo_alta_navegacao)
    sheet.add_data_validation(dv_status_primeira_consula_especialista)
    sheet.add_data_validation(dv_unidade_agendamento)
    sheet.add_data_validation(dv_tipo_tratamento)
    sheet.add_data_validation(dv_cirurgia_solicitada)
    sheet.add_data_validation(dv_unidade_cirurgia)
    
    wb.save(full_file_name)
    wb.close()

# COMMAND ----------

# DBTITLE 1,Função copy_files
def copy_files(source, target):
    source = f"file://{source}"
    # print("-"*120)
    print("Copiando arquivo")
    print(f"De  : {source}")
    print(f"Para: {target}")

    dbutils.fs.cp(source, target, recurse=True)

# COMMAND ----------

# DBTITLE 1,Função extract_data
def extract_data(unidade, ids):
    print("-" * 120)
    print(f"Extraindo dados para a unidade: {unidade}")
    
    _df_export = get_data(ids)

    print(f"Registros encontrados: {len(_df_export.index)}")

    _dic_col_names = get_cols()
    
    _info = export_to_excel(_df_export, _dic_col_names, unidade)
    
    _dic_col_fixed_width = get_cols_config()
    
    format_excel(_info['arquivoLocal'], _dic_col_names, _dic_col_fixed_width)
    
    copy_files(_info['arquivoLocal'], _info['arquivoRemoto'])

    return _info

# COMMAND ----------

# DBTITLE 1,Função clean
def clean(folder):
    files = os.listdir(folder)
    for file in files:
        file_path = os.path.join(folder, file)        
        if os.path.isfile(file_path):
            print(f"Excluido arquivo: {file_path}")
            os.remove(file_path)

# COMMAND ----------

# DBTITLE 1,Carrega as configurações
df = spark.read.option("multiline","true").json(f"{root_remote_config_path}unidades.json")
df.display()

# COMMAND ----------

# DBTITLE 1,Exporta os dados para cada unidade
files = []

for row in df.collect():
    if row.unidades is None:
        print("-" * 120)
        print(f"Unidade: {row.unidade} - Não configurada!")
        continue

    info = extract_data(row.unidade, row.unidades)
    files.append(info)

# COMMAND ----------

# DBTITLE 1,Salva metadados dos arquivos gerados
metadados = f"{current_folder}{id_projeto}_metadados_envio.json"

json_data = json.dumps(files)

with open(metadados, "w") as file:
    file.write(json_data)

# COMMAND ----------

# DBTITLE 1,Copia arquivo com metadados para o storage
copy_files(metadados, root_remote_data_path)

# COMMAND ----------

# DBTITLE 1,Apaga arquivos locais (se existirem)
clean(current_folder)

# COMMAND ----------

# DBTITLE 1,Lista conteúdo da pasta local
!ls -lha {current_folder}

# COMMAND ----------

# DBTITLE 1,Remove pasta
!rm -rf {current_folder}

# COMMAND ----------

# DBTITLE 1,Classe dataClass
@dataclass
class Info:
    fileName: str = ""
    sourcePath: str = ""
    targetPath: str = ""
    environment: str = ""
    unidade: str = ""
    onedriveUser: str = ""
    logicAppUrl: str = ""
    emailToDefault: str = ""
    emailTo: str = ""
    emailCc: str = ""
    emailBcc: str = ""
    emailSubject: str = ""
    emailBody: str = ""
    registros: str = ""
    dataProcessamento: str = ""
    dataProcessamentoFormatada: str = ""

# COMMAND ----------

# MAGIC %md
# MAGIC #Funções

# COMMAND ----------

# DBTITLE 1,load_metadata
def load_metadata(file_path):
    """
    Carrega metadados de um arquivo JSON e cria uma lista de objetos Info.

    Args:
        file_path (str): O caminho do arquivo JSON contendo os metadados.

    Returns:
        list: Uma lista de objetos Info preenchidos com os dados do arquivo JSON.
    """
    info_list = []

    df = spark.read.option("multiline", "true").json(file_path)

    for row in df.collect():
        info_list.append(
            Info(
                environment=environment,
                fileName=row["nomeArquivo"],
                sourcePath=row["caminhoRemoto"],
                unidade=row["unidade"],
                registros=row["registros"],
                dataProcessamento=row["dataProcessamento"],
                dataProcessamentoFormatada=row["dataProcessamentoFormatada"],
            )
        )

    return info_list

# COMMAND ----------

# DBTITLE 1,load_config
def load_config(file_path, info_list, target_path_suffix):
    """
    Carrega a configuração de um arquivo JSON e atualiza a lista de objetos Info.

    Args:
        file_path (str): O caminho do arquivo JSON contendo a configuração.
        info_list (list): Uma lista de objetos Info a serem atualizados.

    Returns:
        list: A lista de objetos Info atualizada com os dados da configuração.
    """
    df = spark.read.option("multiline", "true").json(file_path)
    display(df)

    for row in df.collect():
        for info in info_list:
            info.targetPath = f"{row['targetPath']}{target_path_suffix}"
            info.onedriveUser = row["onedriveUser"]
            info.logicAppUrl = row["logicAppUrl"]
            info.emailToDefault = row["emailToDefault"]

            _subject = "emailSubject" if info.registros > 0 else "emailSubjectNoRecords"
            _body = "emailBody" if info.registros > 0 else "emailBodyNoRecords"

            info.emailSubject = f"{'[' + environment.upper() + ']' if environment != 'prd' else ''}{row[_subject]}"
            info.emailBody = row[_body]

        break

    return info_list

# COMMAND ----------

# DBTITLE 1,load_unidade
def load_unidades(file_path, info_list):
    """
    Carrega informações de unidades de um arquivo JSON e atualiza a lista de objetos Info.

    Args:
        file_path (str): O caminho do arquivo JSON contendo as informações das unidades.
        info_list (list): Uma lista de objetos Info a serem atualizados.

    Returns:
        list: A lista de objetos Info atualizada com os dados das unidades.
    """
    df = spark.read.option("multiline", "true").json(file_path)

    for info in info_list:
        df_filtered = df.filter(df.unidade == info.unidade)

        # if not df_filtered.rdd.isEmpty():
        if df_filtered.count() > 0:
            row = df_filtered.collect()[0]

            info.emailTo = row["emailTo"]
            info.emailCc = row["emailCc"]
            info.emailBcc = row["emailBcc"]
        else:
            info.emailTo = info.emailToDefault

    return info_list

# COMMAND ----------

# DBTITLE 1,send_to_onedrive
def send_to_onedrive(info_list):
    """
    Envia informações para o OneDrive usando a URL da Logic App.

    Args:
        info_list (list): Uma lista de objetos Info contendo as informações a serem enviadas.

    Returns:
        None
    """
    for info in info_list:
        print("-" * 120)
        print(f"Enviando: {info.unidade} - Registros: {info.registros}")
        payload = json.dumps(asdict(info))

        response = requests.post(
            info.logicAppUrl, data=payload, headers={"Content-Type": "application/json"}
        )

# COMMAND ----------

# MAGIC %md
# MAGIC #Carga das configurações

# COMMAND ----------

# DBTITLE 1,Cerrega os metadados
file_path = f"/mnt/trusted/datalake/{catalog_name}/projetos/{id_projeto}/data/{environment}/envio/{id_projeto}_metadados_envio.json"
info_list = load_metadata(file_path)
print(info_list)

# COMMAND ----------

# DBTITLE 1,Carrega as configuracoes de envio
config_path = f"/mnt/trusted/datalake/{catalog_name}/projetos/{id_projeto}/config/{environment}/config_v2.json"
info_list = load_config(config_path, info_list, "Envio/")
print(info_list)

# COMMAND ----------

# DBTITLE 1,Carrega as configuracoes das unidades
unidades_path = f"/mnt/trusted/datalake/{catalog_name}/projetos/{id_projeto}/config/{environment}/unidades.json"
info_list = load_unidades(unidades_path, info_list)
print(info_list)

# COMMAND ----------

# MAGIC %md
# MAGIC #Envia arquivos

# COMMAND ----------

# DBTITLE 1,Envio para o onedrive
send_to_onedrive(info_list)

# COMMAND ----------

# DBTITLE 1,Fim da execucao
dbutils.notebook().exit("Fim da execução!")
