# Databricks notebook source
# MAGIC %md
# MAGIC # Instalação de libs

# COMMAND ----------

!pip install -q openpyxl xlsxwriter unidecode

# COMMAND ----------

# MAGIC %md
# MAGIC # Import libs

# COMMAND ----------

import copy
import json
import openpyxl
import os
import pandas as pd
import subprocess

from datetime import date, datetime
from pathlib import Path

# COMMAND ----------

# MAGIC %md
# MAGIC # Parâmetros do notebook

# COMMAND ----------

# dbutils.widgets.removeAll()

# COMMAND ----------

# DBTITLE 1,Ambiente
dbutils.widgets.text("environment", "dev", "Environment")
environment = dbutils.widgets.get("environment")
print("Environment:", environment)

# COMMAND ----------

# DBTITLE 1,ID Projeto
dbutils.widgets.text("id_projeto", "endometriose", "ID Projeto")
id_projeto = dbutils.widgets.get("id_projeto")
print("id_projeto:", id_projeto)

# COMMAND ----------

# DBTITLE 1,Data Envio Navegação
dbutils.widgets.text("data_enviado_navegacao", "", "Data Enviado Navegação")
data_enviado_navegacao = dbutils.widgets.get("data_enviado_navegacao")
if data_enviado_navegacao == "":
    data_enviado_navegacao = datetime.now().strftime("%Y-%m-%d")

print(f"Data Enviado Navegação: {data_enviado_navegacao}")

# COMMAND ----------

# DBTITLE 1,Define variável que será injetada no nome das tabelas
if environment in ["hml", "prd"]:
    environment_tbl = ""
else:
    environment_tbl = f"{environment}_"

print("environment_tbl:", environment_tbl)

# COMMAND ----------

# DBTITLE 1,Define variáveis com nomes das tabelas
tbl_navegacao = f"{environment_tbl}tbl_gold_modelo_{id_projeto}_navegacao"
vw_analise_retorno = f"{environment_tbl}vw_gold_modelo_{id_projeto}_analise_retorno"

print(tbl_navegacao)
print(vw_analise_retorno)

# COMMAND ----------

# DBTITLE 1,Obtém o caminho da pasta atual
# shell_command = "pwd"
# output = subprocess.check_output(shell_command, shell=True)
# current_folder = os.path.join(output.decode().strip(), f"extracao-navegacao-{id_projeto}") + "/"

current_folder = os.path.join("/tmp", f"extracao-navegacao-{id_projeto}") + "/"

Path(current_folder).mkdir(parents=True, exist_ok=True)

print(current_folder)

# COMMAND ----------

# DBTITLE 1,Define o caminho principal da pasta remota
# root_remote_path = f"/mnt/trusted/datalake/ia/projetos/{id_projeto}/data/{environment}/navegacao/"
# print(root_remote_path)

# COMMAND ----------

# DBTITLE 1,Define o caminho mensal da pasta remota
# year, month, day = data_enviado_navegacao.split("-")
# remote_path = f"{root_remote_path}{year}/{month}/"
# print(remote_path)

# COMMAND ----------

# DBTITLE 1,Define o caminho principal da pasta remota
root_remote_path = f"/mnt/trusted/datalake/ia/projetos/{id_projeto}/"
print(root_remote_path)

# COMMAND ----------

# DBTITLE 1,Define o caminho principal da pasta remota de configs
root_remote_config_path = f"{root_remote_path}config/{environment}/"
print(root_remote_config_path)

# COMMAND ----------

# DBTITLE 1,Define o caminho principal da pasta remota dos dados
root_remote_data_path = f"{root_remote_path}data/{environment}/navegacao/"
print(root_remote_data_path)

# COMMAND ----------

# DBTITLE 1,Define o caminho mensal da pasta remota
year, month, day = data_enviado_navegacao.split("-")
remote_path = f"{root_remote_data_path}{year}/{month}/"
print(remote_path)

# COMMAND ----------

# MAGIC %md
# MAGIC # Dados

# COMMAND ----------

# DBTITLE 1,Exclui registros inseridos na mesma dataEnviadoNavegacao
spark.sql(f"""
    delete from ia.{tbl_navegacao}
    where dataEnviadoNavegacao = '{data_enviado_navegacao}'
""").display()

# COMMAND ----------

# DBTITLE 1,Insere registros na tabela histórica com dados de navegação
spark.sql(f"""
    insert into ia.{tbl_navegacao}
    (
        idNavegacao,
        dataEnviadoNavegacao,
        dataExecucaoModelo,
        idPredicao
    )
    select
        uuid() as idNavegacao,
        '{data_enviado_navegacao}' as dataEnviadoNavegacao,
        dataExecucaoModelo,
        idPredicao
    from ia.{vw_analise_retorno}
    where idRetorno is not null
    and destino = '1 - Navegação'
    and idPredicao not in ( select idPredicao from ia.{tbl_navegacao} )
""").display()

# COMMAND ----------

# MAGIC %md
# MAGIC # Exporta os dados para Excel

# COMMAND ----------

# DBTITLE 1,Função get_data
def get_data(unidades):
  df = spark.sql(f"""
      select
          nav.dataExecucaoModelo,
          nav.idPredicao,
          nav.idExame,
          date_format(date(ret.dataExame), 'dd/MM/yyyy') as dataExame,
          date_format(date(ret.dataLiberacaoLaudo), 'dd/MM/yyyy') as dataLiberacaoLaudo,
          ret.nomeConvenio,
          -- ret.tipoCodigo,
          ret.codigo,
          ret.nomeCodigo,
          ret.descricaoProcedimento,
          -- ret.idUnidade,
          ret.nomeUnidade,
          ret.cnpjUnidade,
          ret.regionalUnidade,
          -- ret.tipoUnidade,
          -- ret.idMedico,
          ret.numCrm,
          ret.ufCrm,
          ret.nomeMedico,
          -- ret.idPaciente,
          ret.cpfPaciente,
          ret.nomePaciente,
          -- ret.dataNascimentoPaciente,
          ret.idadePaciente,
          ret.sexoPaciente,
          ret.telefoneContato,
          ret.laudoOriginal,
          -- round(scoreProbabilidade, 2) as score,
          ret.classificacao,
          ret.achado,
          ret.observacoes,
          ret.destino
          -- "" as observacoes

      from ia.{vw_analise_retorno} as ret

      inner join ia.{tbl_navegacao} as nav
          on ret.idPredicao = nav.idPredicao

      where nav.dataEnviadoNavegacao = '{data_enviado_navegacao}'
          and (idUnidade in ({unidades})
          or "{unidades}" = "'todas'")

      order by
          regionalUnidade,
          nomeUnidade,
          nomePaciente
  """).toPandas()

  return df

# COMMAND ----------

# DBTITLE 1,Função get_cols
def get_cols():
    """Dicionário de/para com os nomes das colunas de saída"""
    dic_col_names = {
        "dataExecucaoModelo": "dataExecucaoModelo",
        "idPredicao": "idPredicao",
        "idExame": "idExame",

        "nomePaciente": "Nome Paciente",
        "cpfPaciente": "CPF Paciente",
        "idadePaciente": "Idade Paciente",
        "sexoPaciente": "Sexo Paciente",
        "telefoneContato": "Telefone Contato",

        "nomeUnidade": "Unidade",
        "cnpjUnidade": "CNPJ Unidade",
        "regionalUnidade": "Regional Unidade",

        "numCrm": "CRM",
        "ufCrm": "UF CRM",
        "nomeMedico": "Nome Médico",

        "dataExame": "Data Exame",
        "dataLiberacaoLaudo": "Data Liberação Laudo",
        "nomeConvenio": "Convênio",
        "codigo": "Código TUSS",
        "nomeCodigo": "Descrição TUSS",
        "descricaoProcedimento": "Nome Procedimento",
        "laudoOriginal": "Laudo",
        "achado": "Achado",
        "observacoes": "Observações",
        "destino": "Destino"
    }

    return dic_col_names

# COMMAND ----------

from datetime import datetime

def extract_date(date_str, date_format='%d/%m/%Y'):
    return datetime.strptime(date_str, date_format).date()

# COMMAND ----------

# DBTITLE 1,Função export_to_excel
def export_to_excel(df, dic_col_names, unidade):
    file_name = f"{id_projeto.upper()}_{unidade}_{data_enviado_navegacao}.xlsx"
    local_file = f"{current_folder}{file_name}"
    remote_file = f"{remote_path}{file_name}"

    cols = [col for col in dic_col_names.keys()]
    df[cols].to_excel(local_file, index=False, freeze_panes=(1, 4), engine='xlsxwriter')

    return {
        "unidade": unidade,
        "nomeArquivo": file_name,
        "arquivoLocal": local_file,
        "arquivoRemoto": remote_file,
        "caminhoRemoto": remote_path.replace("/mnt", ""),
        "registros": len(df.index),
        "dataProcessamento": data_enviado_navegacao,
        "dataProcessamentoFormatada": datetime.strptime(data_enviado_navegacao, "%Y-%m-%d").strftime("%d/%m/%Y"),
    }

# COMMAND ----------

# DBTITLE 1,Função get_cols_config
def get_cols_config():
    """Define tamanho fixo para algumas colunas. ( -1 = Coluna oculta )"""
    dic_col_fixed_width = {
        "dataExecucaoModelo": -1,
        "idPredicao": -1,
        "idExame": -1,
        "Laudo": 150,
        "Achado": 30,
        "Observações": 30,
    }
    return dic_col_fixed_width

# COMMAND ----------

# DBTITLE 1,Função format_excel
def format_excel(full_file_name, dic_col_names, dic_col_fixed_width):
    import math
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.load_workbook(full_file_name)
    sheet = wb.active

    dic_col_size = {}

    # dv_achado = DataValidation(type="list", formula1='"1 - Alta probabilidade,2 - Média probabilidade,3 - Baixa probabilidade,4 - Descartada probabilidade"', allow_blank=True)
    # dv_achado.showInputMessage = True
    # dv_achado.showErrorMessage = True
    # dv_achado.error = "Este valor não corresponde às restrições de valições de dados definidas para esta célula."
    # dv_achado.errorTitle = 'Valor Inválido'

    # dv_linha_cuidado = DataValidation(type="list", formula1='"1 - Cirrose,2 - Esteatose,3 - Transplante"', allow_blank=True)
    # dv_linha_cuidado.showInputMessage = True
    # dv_linha_cuidado.showErrorMessage = True
    # dv_linha_cuidado.error = "Este valor não corresponde às restrições de valições de dados definidas para esta célula."
    # dv_linha_cuidado.errorTitle = 'Valor Inválido'

    # col_idade = ""

    wrap_text = [
        "Laudo",
        # "Tipo Exame",
        # "Nome Hospital",
    ]

    h_align_center = [
        "CPF Paciente",
        "Idade Paciente",
        "Sexo Paciente",
        "Telefone Contato",
        "CNPJ Unidade",
        "Regional Unidade"
        "CRM",
        "UF CRM",
        "Data Liberação Laudo",
        "Código TUSS",
    ]

    for cols in sheet.iter_cols():
        for cell in cols:
            # Renomeia as colunas
            if cell.row == 1:
                cell.value = dic_col_names.get(cell.value, cell.value)

            column_letter = cell.column_letter
            len_value = len(str(cell.value))
            
            if len_value > dic_col_size.get(column_letter, 0):
                dic_col_size[column_letter] = len_value

            # if cols[0].value == "Achado" and cell.row > 1:
            #     dv_achado.add(cell)

            # if cols[0].value == "Linha de Cuidado" and cell.row > 1:
            #     dv_linha_cuidado.add(cell)

            # if cols[0].value == "Idade Paciente":
            #     if cell.row > 1:
            #         idade = cell.value
            #         if idade > IDADE_LIMITE:
            #             cell.fill = PatternFill(fill_type="solid", start_color="F2DCDB")
            #             cell.font = Font(color="9C0006")

            alignment = copy.copy(cell.alignment)
            alignment.vertical = "center"

            if cell.row > 1 and cols[0].value in wrap_text:
                alignment.wrapText = True

            if cell.row > 1 and cols[0].value in h_align_center:
                alignment.horizontal = "center"
        
            cell.alignment = alignment

    # Ajusta a largura das colunas
    for k, v in dic_col_size.items():
        adjusted_width = dic_col_fixed_width.get(sheet[f"{k}1"].value)

        if adjusted_width is None:
            adjusted_width = int((v + 2))

        if adjusted_width == -1:
            sheet.column_dimensions[k].hidden = True
        else:
            sheet.column_dimensions[k].width = adjusted_width
                
    # sheet.add_data_validation(dv_achado)
    # sheet.add_data_validation(dv_linha_cuidado)

    wb.save(full_file_name)
    wb.close()


# COMMAND ----------

# DBTITLE 1,Função copy_files
def copy_files(source, target):
    source = f"file://{source}"
    # print("-"*120)
    print("Copiando arquivo")
    print(f"De  : {source}")
    print(f"Para: {target}")

    dbutils.fs.cp(source, target, recurse=True)

# COMMAND ----------

# DBTITLE 1,Função extract_data
def extract_data(unidade, ids):
    print("-" * 120)
    print(f"Extraindo dados para a unidade: {unidade}")
    
    _df_export = get_data(ids)

    print(f"Registros encontrados: {len(_df_export.index)}")

    _dic_col_names = get_cols()
    
    _info = export_to_excel(_df_export, _dic_col_names, unidade)
    
    _dic_col_fixed_width = get_cols_config()
    
    format_excel(_info['arquivoLocal'], _dic_col_names, _dic_col_fixed_width)
    
    copy_files(_info['arquivoLocal'], _info['arquivoRemoto'])

    return _info

# COMMAND ----------

# DBTITLE 1,Função clean
def clean(folder):
    files = os.listdir(folder)
    for file in files:
        file_path = os.path.join(folder, file)        
        if os.path.isfile(file_path):
            print(f"Excluido arquivo: {file_path}")
            os.remove(file_path)

# COMMAND ----------

# DBTITLE 1,Carrega as configurações
df = spark.read.option("multiline","true").json(f"{root_remote_config_path}navegacao.json")
df.display()

# COMMAND ----------

# DBTITLE 1,Exporta os dados para cada unidade
files = []

for row in df.collect():
    if row.unidades is None:
        print("-" * 120)
        print(f"Unidade: {row.unidade} - Não configurada!")
        continue

    info = extract_data(row.unidade, row.unidades)
    files.append(info)

# COMMAND ----------

# DBTITLE 1,Salva metadados dos arquivos gerados
metadados = f"{current_folder}{id_projeto}_metadados_navegacao.json"

json_data = json.dumps(files)

with open(metadados, "w") as file:
    file.write(json_data)

# COMMAND ----------

# DBTITLE 1,Copia arquivo com metadados para o storage
copy_files(metadados, root_remote_data_path)

# COMMAND ----------

# DBTITLE 1,Apaga arquivos locais (se existirem)
clean(current_folder)

# COMMAND ----------

# DBTITLE 1,Lista conteúdo da pasta local
!ls -lha {current_folder}

# COMMAND ----------

# DBTITLE 1,Remove pasta
!rm -rf {current_folder}

# COMMAND ----------

# DBTITLE 1,Fim do processamento!
dbutils.notebook.exit("Fim do processamento!")

# COMMAND ----------


