# Databricks notebook source
# MAGIC %md
# MAGIC # Instalação de libs

# COMMAND ----------

# DBTITLE 1,pip install
!pip install -q openpyxl xlsxwriter

# COMMAND ----------

# DBTITLE 1,restartPython
dbutils.library.restartPython()

# COMMAND ----------

# MAGIC %md
# MAGIC # Imports

# COMMAND ----------

# DBTITLE 1,Import
import os
import pandas as pd
import openpyxl
import copy
import json
import requests
import subprocess

from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path

# COMMAND ----------

# DBTITLE 1,pd.set_option
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.max_colwidth', 230)

# COMMAND ----------

# MAGIC %md
# MAGIC # Parâmetros

# COMMAND ----------

# dbutils.widgets.removeAll()

# COMMAND ----------

# DBTITLE 1,ID Projeto
dbutils.widgets.text("id_projeto", "hepatologia", "ID Projeto")
id_projeto = dbutils.widgets.get("id_projeto")
print("id_projeto:", id_projeto)

# COMMAND ----------

# DBTITLE 1,Ambiente
dbutils.widgets.text("environment", "dev", "Environment")
environment = dbutils.widgets.get("environment")
print("environment:", environment)

# COMMAND ----------

# DBTITLE 1,Prefixo Tabelas
environment_tbl = "" if environment in ["hml", "prd"] else f"{environment}_"
print("environment_tbl:", environment_tbl)

# COMMAND ----------

# DBTITLE 1,Catalog
dbutils.widgets.text("catalog", "diamond_hepatologia", "Catalog")
catalog_name = dbutils.widgets.get("catalog")
print(f"catalog_name: {catalog_name}")

# COMMAND ----------

# DBTITLE 1,Schema
dbutils.widgets.text("schema", "hepatologia", "Schema")
schema_name = dbutils.widgets.get("schema")
print(f"schema_name: {schema_name}")

# COMMAND ----------

# DBTITLE 1,Work Catalog
dbutils.widgets.text("work_catalog", "diamond_hepatologia", "Work Catalog")
work_catalog_name = dbutils.widgets.get("work_catalog")
print(f"work_catalog_name: {work_catalog_name}")

# COMMAND ----------

# DBTITLE 1,Work Schema
dbutils.widgets.text("work_schema", "workarea", "Work Schema")
work_schema_name = dbutils.widgets.get("work_schema")
print(f"work_schema_name: {work_schema_name}")

# COMMAND ----------

# DBTITLE 1,Data Execução
dbutils.widgets.text("data_execucao_modelo", "", "Data Execução Modelo")
data_execucao_modelo = dbutils.widgets.get("data_execucao_modelo")
if data_execucao_modelo == "":
    data_execucao_modelo = datetime.now().strftime("%Y-%m-%d")
print(f"Data Referencia: {data_execucao_modelo}")

# COMMAND ----------

# DBTITLE 1,Main Catalog
if environment in ["hml", "prd"]:
    main_catalog = catalog_name + ("" if schema_name == "" else f".{schema_name}")
else:
    main_catalog = work_catalog_name + ("" if work_schema_name == "" else f".{work_schema_name}")

print(f"main_catalog: {main_catalog}")

# COMMAND ----------

# DBTITLE 1,Work Catalog
work_catalog = work_catalog_name + ("" if work_schema_name == "" else f".{work_schema_name}")
print(f"work_catalog: {work_catalog}")

# COMMAND ----------

# DBTITLE 1,Pasta raiz onde os dados devem ser salvos
if schema_name == "":
    root_folder = f"/mnt/trusted/datalake/{main_catalog}/data/{id_projeto}/diamond"
else:
    root_folder = f"abfss://artificial-intelligence@sardslusdevelopmenthml.dfs.core.windows.net/curated/ia/diamond/{id_projeto}/{environment}"

print(root_folder)

# COMMAND ----------

# DBTITLE 1,Define o caminho principal da pasta remota
root_remote_path = f"/mnt/trusted/datalake/ia/projetos/{id_projeto}/"
print(root_remote_path)

# COMMAND ----------

# DBTITLE 1,Define o caminho principal da pasta remota de configs
root_remote_config_path = f"{root_remote_path}config/{environment}/"
print(root_remote_config_path)

# COMMAND ----------

# DBTITLE 1,Define o caminho principal da pasta remota dos dados
root_remote_data_path = f"{root_remote_path}data/{environment}/envio/"
print(root_remote_data_path)

# COMMAND ----------

# DBTITLE 1,Define o caminho mensal da pasta remota
year, month, day = data_execucao_modelo.split("-")
remote_path = f"{root_remote_data_path}{year}/{month}/"
print(remote_path)

# COMMAND ----------

# DBTITLE 1,Obtém o caminho da pasta atual
current_folder = os.path.join("/tmp", id_projeto) + "/"

Path(current_folder).mkdir(parents=True, exist_ok=True)

print(current_folder)

# COMMAND ----------

# DBTITLE 1,Exibe o conteúdo da pasta atual
ls -lha {current_folder}

# COMMAND ----------

# DBTITLE 1,Constantes
IDADE_LIMITE = 75

# COMMAND ----------

# MAGIC %md
# MAGIC # Funções Auxiliares

# COMMAND ----------

# DBTITLE 1,optimize_table
def optimize_table(table_id):
    spark.sql(f"VACUUM {table_id}")
    spark.sql(f"OPTIMIZE {table_id}")
    spark.sql(f"ANALYZE TABLE {table_id} COMPUTE STATISTICS")

# COMMAND ----------

# DBTITLE 1,table_location
table_location = lambda x: f"{root_folder}/{x.split('.')[-1]}"

# COMMAND ----------

# MAGIC %md
# MAGIC # Variáveis com os nomes das tabelas

# COMMAND ----------

# DBTITLE 1,Define variáveis com nomes das tabelas
tbl_entrada = f"{main_catalog}.{environment_tbl}tb_diamond_mod_hepatologia_entrada"
tbl_saida = f"{main_catalog}.{environment_tbl}tb_diamond_mod_hepatologia_saida"
vw_name_saida = f"{main_catalog}.{environment_tbl}vw_diamond_mod_hepatologia"

# COMMAND ----------

# DBTITLE 1,Exibe valores das variáveis com nomes das tabelas
print(f"{'tbl_entrada':15}: {tbl_entrada}")
print(f"{'tbl_saida':15}: {tbl_saida}")
print(f"{'vw_name_saida':15}: {vw_name_saida}")

# COMMAND ----------

# MAGIC %md
# MAGIC # Exporta os dados para Excel

# COMMAND ----------

vw_name_saida

# COMMAND ----------

# DBTITLE 1,Função get_data
def get_data(unidades):
  df = spark.sql(f"""
    with cte_input as (
        select
            dt_execucao as dataExecucaoModelo,
            id_predicao as idPredicao,
            id_exame as idExame,
            nome_paciente as nomePaciente,
            dt_nascimento_paciente as dataNascimentoPaciente,
            num_idade_paciente as idadePaciente,
            nome_convenio as nomeConvenio,
            id_unidade as idHospital,
            emp_nome_unidade as nomeHospital,
            emp_regional_unidade as regionalHospital,
            nome_medico as nomeMedico,
            doc_crm_medico as numCrm,
            uf_crm_medico as ufCrm,
            dt_exame as dataExame,
            proced_nome_exame as tipoExame,
            proced_laudo_exame_original as laudoExameOriginal,
            vl_plt as valorPlt,
            num_dif_tempo_imagem_plt as difTempoImagemPlt,
            fl_relevante as flgRelevante
        from {vw_name_saida}
        where dt_execucao = date('{data_execucao_modelo}')
          and fl_relevante = true
          and id_unidade in ({unidades})
          and nome_convenio not rlike r'(?i)(?<!\w)AMIL(?:\s*[-/().A-Za-z0-9]+)?'
    )
    ,cte as (
        select
            dataExecucaoModelo,
            idPredicao,
            idExame,
            nomePaciente,
            ifnull(idadePaciente, cast(floor(date_diff(day, dataNascimentoPaciente, current_date()) / 365.25) as int)) as idadePaciente,
            nomeHospital,
            regionalHospital,

            nomeMedico,
            numCrm,
            ufCrm,

            date_format(date(dataExame), 'dd/MM/yyyy') as dataExame,
            trim(REGEXP_REPLACE(tipoExame, '[^A-Za-z0-9\s\.,;!?\-]+', '')) as tipoExame,
            laudoExameOriginal as laudoExame,
            valorPlt,
            difTempoImagemPlt,
            "" as achadoRelevante,
            "" as achado,
            "" as linhaCuidado,
            "" as observacao,
            "" as destino,
            "" as prioridade,

            case 
                when trim(upper(regionalHospital)) = 'RJ' then 1
                when trim(upper(regionalHospital)) = 'SP' then 2
                when trim(upper(regionalHospital)) = 'PB' then 3
                else 4
            end as regionalHospitalOrder
            
        from cte_input
    )
    select * except(regionalHospitalOrder)
    from cte
    where idadePaciente <= 80
    order by
        regionalHospitalOrder,
        nomePaciente,
        dataExame
  """).toPandas()

  return df

# COMMAND ----------

# DBTITLE 1,Função get_cols
def get_cols():
    """Dicionário de/para com os nomes das colunas de saída"""
    dic_col_names = {
        "dataExecucaoModelo": "Data Referência",
        "idPredicao": "idPredicao",
        "idExame": "idExame",
        "nomePaciente": "Nome Paciente",
        "idadePaciente": "Idade Paciente",
        "nomeHospital": "Nome Hospital",
        "regionalHospital": "UF Hospital",
        "nomeMedico": "Médico Solicitante",
        "numCrm": "CRM",
        "ufCrm": "UF CRM",
        "dataExame": "Data Exame",
        "tipoExame": "Tipo Exame",
        "laudoExame": "Laudo",
        "valorPlt": "Valor PLT",
        "difTempoImagemPlt": "Dif. Tempo Imagem PLT",
        "achadoRelevante": "Achado Relevante",
        "achado": "Achado",
        "linhaCuidado": "Linha de Cuidado",
        "destino": "Destino",
        "prioridade": "Prioridade",
        "observacao": "Observação",
    }

    return dic_col_names

# COMMAND ----------

# DBTITLE 1,Função export_to_excel
def export_to_excel(df, dic_col_names, unidade):
    file_name = f"{id_projeto}_{unidade}_{data_execucao_modelo}.xlsx"
    local_file = f"{current_folder}{file_name}"
    remote_file = f"{remote_path}{file_name}"

    cols = [col for col in dic_col_names.keys()]
    df[cols].to_excel(local_file, index=False, freeze_panes=(1, 4), engine='xlsxwriter')

    return {
        "unidade": unidade,
        "nomeArquivo": file_name,
        "arquivoLocal": local_file,
        "arquivoRemoto": remote_file,
        "caminhoRemoto": remote_path.replace("/mnt", ""),
        "registros": len(df.index),
        "dataProcessamento": data_execucao_modelo,
        "dataProcessamentoFormatada": datetime.strptime(data_execucao_modelo, "%Y-%m-%d").strftime("%d/%m/%Y"),
    }

# COMMAND ----------

# DBTITLE 1,Função get_cols_config
def get_cols_config():
    """Define tamanho fixo para algumas colunas. ( -1 = Coluna oculta )"""
    dic_col_fixed_width = {
        "Data Referência": -1,
        "idPredicao": -1,
        "idExame": -1,
        "Nome Hospital": 30,
        "Tipo Exame": 15,
        "Laudo": 150,
        "Achado Relevante": 37,
        "Achado": 80,
        "Linha de Cuidado": 27,
        "Destino": 80,
        "Observação": 80,
    }
    return dic_col_fixed_width

# COMMAND ----------

# DBTITLE 1,Função format_excel
def format_excel(full_file_name, dic_col_names, dic_col_fixed_width):
    import math
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.load_workbook(full_file_name)
    sheet = wb.active

    dic_col_size = {}

    data_validation_error_title = "Valor Inválido"
    data_validation_error_message = "Este valor não corresponde às restrições de valições de dados definidas para esta célula."
    
    dv_achado = DataValidation(type="list", formula1='"1 - Sim (Tem Doença Fígado),2 - Sim (Mas Não Tem Doença Fígado),3 - Não"', allow_blank=True)
    dv_achado.showInputMessage = True
    dv_achado.showErrorMessage = True
    dv_achado.error = data_validation_error_message
    dv_achado.errorTitle = data_validation_error_title

    dv_linha_cuidado = DataValidation(type="list", formula1='"1 - Cirrose,2 - Esteatose,3 - Transplante,4 - Cirurgia Hepatobiliar Pancreática,5 - Nódulo Benigno,6 - Outros"', allow_blank=True)
    dv_linha_cuidado.showInputMessage = True
    dv_linha_cuidado.showErrorMessage = True
    dv_linha_cuidado.error = data_validation_error_message
    dv_linha_cuidado.errorTitle = data_validation_error_title

    dv_navegacao = DataValidation(type="list", formula1='"1 - Navegação Transplante,2 - Navegação Origem Gastro,3 - Navegação Origem Hepato,4 - Navegação Outros,5 - Navegação Oncologia,6 - Não"', allow_blank=True)
    dv_navegacao.showInputMessage = True
    dv_navegacao.showErrorMessage = True
    dv_navegacao.error = data_validation_error_message
    dv_navegacao.errorTitle = data_validation_error_title

    dv_prioridade = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
    dv_prioridade.showInputMessage = True
    dv_prioridade.showErrorMessage = True
    dv_prioridade.error = data_validation_error_message
    dv_prioridade.errorTitle = data_validation_error_title

    # col_idade = ""

    wrap_text = [
        "Laudo",
        "Tipo Exame",
        "Nome Hospital",
    ]

    h_align = [
        "Idade Paciente",
        "UF Hospital",
        "Data Exame",
        "Valor PLT",
        "Dif. Tempo Imagem PLT",
        "Achado Relevante",
        "Linha de Cuidado",
        "CRM",
        "UF CRM",
        "Prioridade"
    ]

    for cols in sheet.iter_cols():
        for cell in cols:
            # Renomeia as colunas
            if cell.row == 1:
                cell.value = dic_col_names.get(cell.value, cell.value)

            column_letter = cell.column_letter
            len_value = len(str(cell.value))
            
            if len_value > dic_col_size.get(column_letter, 0):
                dic_col_size[column_letter] = len_value

            if cols[0].value == "Achado Relevante" and cell.row > 1:
                dv_achado.add(cell)

            if cols[0].value == "Linha de Cuidado" and cell.row > 1:
                dv_linha_cuidado.add(cell)

            if cols[0].value == "Destino" and cell.row > 1:
                dv_navegacao.add(cell)

            if cols[0].value == "Prioridade" and cell.row > 1:
                dv_prioridade.add(cell)

            if cols[0].value == "Idade Paciente":
                if cell.row > 1:
                    idade = cell.value
                    if idade > IDADE_LIMITE:
                        cell.fill = PatternFill(fill_type="solid", start_color="F2DCDB")
                        cell.font = Font(color="9C0006")

            alignment = copy.copy(cell.alignment)
            alignment.vertical = "center"

            if cell.row > 1 and cols[0].value in wrap_text:
                alignment.wrapText = True

            if cell.row > 1 and cols[0].value in h_align:
                alignment.horizontal = "center"
        
            cell.alignment = alignment

    # Ajusta a largura das colunas
    for k, v in dic_col_size.items():
        adjusted_width = dic_col_fixed_width.get(sheet[f"{k}1"].value)

        if adjusted_width is None:
            adjusted_width = int((v + 2))

        if adjusted_width == -1:
            sheet.column_dimensions[k].hidden = True
        else:
            sheet.column_dimensions[k].width = adjusted_width
                
    sheet.add_data_validation(dv_achado)
    sheet.add_data_validation(dv_linha_cuidado)
    sheet.add_data_validation(dv_navegacao)
    sheet.add_data_validation(dv_prioridade)

    wb.save(full_file_name)
    wb.close()


# COMMAND ----------

# DBTITLE 1,Função copy_files
def copy_files(source, target):
    source = f"file://{source}"
    # print("-"*120)
    print("Copiando arquivo")
    print(f"De  : {source}")
    print(f"Para: {target}")

    dbutils.fs.cp(source, target, recurse=True)

# COMMAND ----------

# DBTITLE 1,Função extract_data
def extract_data(unidade, ids):
    print("-" * 120)
    print(f"Extraindo dados para a unidade: {unidade}")
    
    _df_export = get_data(ids)

    print(f"Registros encontrados: {len(_df_export.index)}")

    _dic_col_names = get_cols()
    
    _info = export_to_excel(_df_export, _dic_col_names, unidade)
    
    _dic_col_fixed_width = get_cols_config()
    
    format_excel(_info['arquivoLocal'], _dic_col_names, _dic_col_fixed_width)
    
    copy_files(_info['arquivoLocal'], _info['arquivoRemoto'])

    return _info

# COMMAND ----------

# DBTITLE 1,Função clean
def clean(folder):
    files = os.listdir(folder)
    for file in files:
        file_path = os.path.join(folder, file)        
        if os.path.isfile(file_path):
            print(f"Excluido arquivo: {file_path}")
            os.remove(file_path)

# COMMAND ----------

# DBTITLE 1,Carrega as configurações
df = spark.read.option("multiline","true").json(f"{root_remote_config_path}unidades.json")
df.display()

# COMMAND ----------

# DBTITLE 1,Exporta os dados para cada unidade
files = []

for row in df.collect():
    if row.unidades is None:
        print("-" * 120)
        print(f"Unidade: {row.unidade} - Não configurada!")
        continue

    info = extract_data(row.unidade, row.unidades)
    files.append(info)

# COMMAND ----------

# DBTITLE 1,Salva metadados dos arquivos gerados
metadados = f"{current_folder}{id_projeto}_metadados_envio.json"

json_data = json.dumps(files)

with open(metadados, "w") as file:
    file.write(json_data)

# COMMAND ----------

# DBTITLE 1,Copia arquivo com metadados para o storage
copy_files(metadados, root_remote_data_path)

# COMMAND ----------

# DBTITLE 1,Apaga arquivos locais (se existirem)
clean(current_folder)

# COMMAND ----------

# DBTITLE 1,Lista conteúdo da pasta local
!ls -lha {current_folder}

# COMMAND ----------

# DBTITLE 1,Remove pasta
!rm -rf {current_folder}

# COMMAND ----------

# MAGIC %md
# MAGIC # Envio de arquivos para o OneDrive

# COMMAND ----------

# MAGIC %md
# MAGIC ## Dataclass Info

# COMMAND ----------

@dataclass
class Info:
    fileName: str = ""
    sourcePath: str = ""
    targetPath: str = ""
    environment: str = ""
    unidade: str = ""
    onedriveUser: str = ""
    logicAppUrl: str = ""
    emailToDefault: str = ""
    emailTo: str = ""
    emailCc: str = ""
    emailBcc: str = ""
    emailSubject: str = ""
    emailBody: str = ""
    registros: str = ""
    dataProcessamento: str = ""
    dataProcessamentoFormatada: str = ""

# COMMAND ----------

# MAGIC %md
# MAGIC ## Funções

# COMMAND ----------

# DBTITLE 1,load_metadata
def load_metadata(file_path):
    """
    Carrega metadados de um arquivo JSON e cria uma lista de objetos Info.

    Args:
        file_path (str): O caminho do arquivo JSON contendo os metadados.

    Returns:
        list: Uma lista de objetos Info preenchidos com os dados do arquivo JSON.
    """
    info_list = []

    df = spark.read.option("multiline", "true").json(file_path)

    for row in df.collect():
        info_list.append(
            Info(
                environment=environment,
                fileName=row["nomeArquivo"],
                sourcePath=row["caminhoRemoto"],
                unidade=row["unidade"],
                registros=row["registros"],
                dataProcessamento=row["dataProcessamento"],
                dataProcessamentoFormatada=row["dataProcessamentoFormatada"],
            )
        )

    return info_list

# COMMAND ----------

# DBTITLE 1,load_config
def load_config(file_path, info_list, target_path_suffix):
    """
    Carrega a configuração de um arquivo JSON e atualiza a lista de objetos Info.

    Args:
        file_path (str): O caminho do arquivo JSON contendo a configuração.
        info_list (list): Uma lista de objetos Info a serem atualizados.

    Returns:
        list: A lista de objetos Info atualizada com os dados da configuração.
    """
    df = spark.read.option("multiline", "true").json(file_path)
    display(df)

    for row in df.collect():
        for info in info_list:
            info.targetPath = f"{row['targetPath']}{target_path_suffix}"
            info.onedriveUser = row["onedriveUser"]
            info.logicAppUrl = row["logicAppUrl"]
            info.emailToDefault = row["emailToDefault"]

            _subject = "emailSubject" if info.registros > 0 else "emailSubjectNoRecords"
            _body = "emailBody" if info.registros > 0 else "emailBodyNoRecords"

            info.emailSubject = f"{'[' + environment.upper() + ']' if environment != 'prd' else ''}{row[_subject]}"
            info.emailBody = row[_body]

        break

    return info_list

# COMMAND ----------

# DBTITLE 1,load_unidades
def load_unidades(file_path, info_list):
    """
    Carrega informações de unidades de um arquivo JSON e atualiza a lista de objetos Info.

    Args:
        file_path (str): O caminho do arquivo JSON contendo as informações das unidades.
        info_list (list): Uma lista de objetos Info a serem atualizados.

    Returns:
        list: A lista de objetos Info atualizada com os dados das unidades.
    """
    df = spark.read.option("multiline", "true").json(file_path)

    for info in info_list:
        df_filtered = df.filter(df.unidade == info.unidade)

        # if not df_filtered.rdd.isEmpty():
        if df_filtered.count() > 0:
            row = df_filtered.collect()[0]

            info.emailTo = row["emailTo"]
            info.emailCc = row["emailCc"]
            info.emailBcc = row["emailBcc"]
        else:
            info.emailTo = info.emailToDefault

    return info_list

# COMMAND ----------

# DBTITLE 1,send_to_onedrive
def send_to_onedrive(info_list):
    """
    Envia informações para o OneDrive usando a URL da Logic App.

    Args:
        info_list (list): Uma lista de objetos Info contendo as informações a serem enviadas.

    Returns:
        None
    """
    for info in info_list:
        print("-" * 120)
        print(f"Enviando: {info.unidade} - Registros: {info.registros}")
        payload = json.dumps(asdict(info))

        response = requests.post(
            info.logicAppUrl, data=payload, headers={"Content-Type": "application/json"}
        )

# COMMAND ----------

# MAGIC %md
# MAGIC ## Carga das configurações

# COMMAND ----------

# DBTITLE 1,Carrega os metadados
file_path = f"/mnt/trusted/datalake/ia/projetos/{id_projeto}/data/{environment}/envio/{id_projeto}_metadados_envio.json"
info_list = load_metadata(file_path)
print(info_list)

# COMMAND ----------

# DBTITLE 1,Carrega as configurações de envio
config_path = f"/mnt/trusted/datalake/ia/projetos/{id_projeto}/config/{environment}/config_v2.json"
info_list = load_config(config_path, info_list, "Envio/")
print(info_list)

# COMMAND ----------

# DBTITLE 1,Carrega configurações das Unidades
unidades_path = f"/mnt/trusted/datalake/ia/projetos/{id_projeto}/config/{environment}/unidades.json"
info_list = load_unidades(unidades_path, info_list)
print(info_list)

# COMMAND ----------

# MAGIC %md
# MAGIC ## Envia arquivos

# COMMAND ----------

send_to_onedrive(info_list)

# COMMAND ----------

# MAGIC %md
# MAGIC # Fim da execução

# COMMAND ----------

dbutils.notebook().exit("Fim da execução!")
